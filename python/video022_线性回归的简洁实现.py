import torch
from torch.utils import data
from d2l import torch as d2l
import openpyxl
import os

true_w = torch.tensor([2, -3.4])
true_b = 4.2
features, labels = d2l.synthetic_data(true_w, true_b, 1000)

def load_array(data_arrays, batch_size, is_train=True):
    dataset = data.TensorDataset(*data_arrays)
    return data.DataLoader(dataset, batch_size, shuffle=is_train)
    

batch_size = 10
data_iter = load_array((features, labels), batch_size)

from torch import nn
# 使用 PyTorch 高层 API 构建一个线性回归模型：
# - 输入维度为2（对应两个特征 x1 和 x2）
# - 输出维度为1（标量 y）
# nn.Sequential 这里仅包含一个 nn.Linear 层，等价于 y_hat = X @ W^T + b
net = nn.Sequential(nn.Linear(2, 1))
# 参数初始化：
# - 权重初始化为均值0、标准差0.01的正态分布，有利于加速初期训练并避免梯度过大
# - 偏置初始化为0
net[0].weight.data.normal_(0, 0.01)
net[0].bias.data.fill_(0)

# 均方误差损失（MSE），与从零实现中的平方损失一致（只是未除以2的版本）
loss = nn.MSELoss()
# 随机梯度下降（SGD）优化器，学习率设为0.03
trainer = torch.optim.SGD(net.parameters(), lr=0.03)

# 为了输出中间过程数据，记录每一步（batch）以及每个 epoch 的指标
param_logs = []   # 记录 [epoch, step, w1, w2, b, batch_loss]

num_epochs = 3
for epoch in range(num_epochs):
    # 每个 epoch 开始时将步数计数器置0
    step = 0
    for X, y in data_iter:
        # 前向计算：得到预测值 y_hat
        l = loss(net(X), y)
        # 反向传播前先清零上一步的梯度，避免梯度累积
        trainer.zero_grad()
        # 反向传播：自动计算每个参数的梯度
        l.backward()
        # 参数更新：按照 SGD 规则使用梯度对权重和偏置进行更新
        trainer.step()
        # 步数自增
        step += 1
        # 记录当前步的参数与损失
        # - 线性层的权重形状为 [1, 2]，对应 [w1, w2]
        # - 偏置形状为 [1]，对应标量 b
        w_vec = net[0].weight.data.view(-1)
        b_vec = net[0].bias.data.view(-1)
        w1 = float(w_vec[0])
        w2 = float(w_vec[1])
        b0 = float(b_vec[0])
        batch_loss = float(l.item())
        param_logs.append([epoch + 1, step, w1, w2, b0])
    l = loss(net(features), labels)
    print(f"epoch {epoch + 1}, loss {l:f}")

# 将中间过程与数据集写入 Excel 文件，便于查看与分析
def save_datas(features, labels, param_logs):
    """
    将训练使用的数据与中间过程指标保存到 Excel：
    - features 工作表：两列 x1, x2，包含全部特征数据
    - labels   工作表：一列 y，包含全部标签数据
    - params   工作表：每个 batch 的参数与损失 [epoch, step, w1, w2, b, batch_loss]
    - epochs   工作表：每个 epoch 的总体训练损失 [epoch, train_loss]
    保存路径位于当前脚本父目录下的 output/video022_linear_regression_log.xlsx，
    若 output 目录不存在则自动创建。
    """
    wb = openpyxl.Workbook()
    # 特征工作表
    ws_features = wb.active
    ws_features.title = "features"
    ws_features.append(["x1", "x2"])
    for row in features.tolist():
        ws_features.append(row)
    # 标签工作表
    ws_labels = wb.create_sheet("labels")
    ws_labels.append(["y"])
    for v in labels.view(-1).tolist():
        ws_labels.append([v])
    # 每步参数与损失工作表
    ws_params = wb.create_sheet("params")
    ws_params.append(["epoch", "step", "w1", "w2", "b"])
    for log in param_logs:
        ws_params.append(log)
    # 每个 epoch 的损失工作表
    ws_epochs = wb.create_sheet("epochs")
    ws_epochs.append(["epoch", "train_loss"])
    # 生成保存路径：父目录的 output
    parent_dir = os.path.dirname(os.path.dirname(__file__))
    output_dir = os.path.join(parent_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(output_dir, "video022_linear_regression_log.xlsx")
    wb.save(save_path)

# 执行保存
save_datas(features, labels, param_logs)
